from datetime import datetime
import cv2
import numpy as np
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image
import time
from PIL import ImageGrab
import os
import hashlib

GUILD_TEMPLATE = cv2.imread(r"data\guild.png", cv2.IMREAD_GRAYSCALE)
GUILD_MASK = cv2.imread(r"data\mask.png", cv2.IMREAD_GRAYSCALE)
GUILD_TITLE = cv2.imread(r"data\guild_member_participation.png", cv2.IMREAD_GRAYSCALE)
NUMBERS = cv2.imread(r"data\numbers.png", cv2.IMREAD_GRAYSCALE)
COMMA = cv2.imread(r"data\comma.png", cv2.IMREAD_GRAYSCALE) * 255
THOUSAND = cv2.imread(r"data\number_1000.png", cv2.IMREAD_GRAYSCALE) * 255

GUILD_H, GUILD_W = GUILD_TEMPLATE.shape
NUMBERS = NUMBERS.reshape(8, 10, 5).swapaxes(0, 1) * 255

BLANK_HASH = "961370590ed3d5eb476badfcdfd8c056fd8871c7f9a719c0c1ad10e9ee5431aa"
TITLE_THRESHOLD = 10000
waiting_second = 10

written_nicknames = {}
scanned_query = 0
is_last_page_captured = 0

print("전체 화면을 캡처해 길드 창의 위치를 찾습니다. 모니터의 해상도가 높을수록 오래 걸릴 수 있습니다.")

capture = np.array(ImageGrab.grab(all_screens=True))
capture = cv2.cvtColor(capture, cv2.COLOR_BGR2GRAY)
match_result = cv2.minMaxLoc(cv2.matchTemplate(capture, GUILD_TEMPLATE, cv2.TM_SQDIFF_NORMED, mask=GUILD_MASK))
best_error, worst_error, best_location, worst_location = match_result

print(f"이미지 오차 점수: {best_error:.3f}, 좌상단 픽셀 좌표 (x, y): {best_location}")
if best_error > 0.1:
    input("오차는 대체로 0.01 이하입니다. [Enter] 키를 입력해 무시하거나 다시 시작해 주세요.")

guild_x, guild_y = best_location
capture = capture[guild_y:, guild_x:]

while np.sum(GUILD_TITLE - capture[64:83, 175:286]) > TITLE_THRESHOLD:
    inputs = input("현재 탭이 [길드원 참여 현황]이 아닙니다. [Enter] 키를 입력해 재확인합니다. [SKIP]을 입력해 무시할 수 있습니다.")
    if inputs.upper() == 'SKIP':
        break
    capture = np.array(ImageGrab.grab(bbox=(guild_x, guild_y, guild_x + GUILD_W, guild_y + GUILD_H), all_screens=True))
    capture = cv2.cvtColor(capture, cv2.COLOR_BGR2GRAY)

start_time = time.time()
while True:
    x = guild_x + 205
    y = guild_y + 131
    w = guild_x + 640
    h = guild_y + 539
    capture = np.array(ImageGrab.grab(bbox=(x, y, w, h), all_screens=True))
    # capture.shape = (H, W)

    # text color is one of (255, 179, 187)
    capture = np.prod(capture == 255, axis=-1) + np.prod(capture == 179, axis=-1) + np.prod(capture == 187, axis=-1)
    capture = capture.astype('u1')

    def detect_and_update(score, start, multiplier=1):
        idx, val = np.where(
            np.all(capture[:, start:start + 5].reshape(17, 24, -1)[np.newaxis, :, 9:17] == NUMBERS[:, np.newaxis, ...],
                   axis=(2, 3)).T)
        score[idx] += val * multiplier

    # mission_score is one of (0, 1, 2, 3, 4, 5). But same detecting method applied.
    mission_scores = np.zeros(17, dtype=int)
    detect_and_update(mission_scores, 274)

    # flag_score is 1,000 or XX0
    flag_scores = np.all(capture[:, 408:432].reshape(17, 24, -1)[:, 9:18] == THOUSAND[np.newaxis, :, :], axis=(1, 2)) * 1000
    detect_and_update(flag_scores, 410, 100)
    detect_and_update(flag_scores, 416, 10)
    # flag_score has no digit for 1

    canal_scores = np.zeros(17, dtype=int)

    detect_and_update(canal_scores, 325, 10000)
    detect_and_update(canal_scores, 331, 1000)
    detect_and_update(canal_scores, 340, 100)
    detect_and_update(canal_scores, 346, 10)
    detect_and_update(canal_scores, 352, 1)

    detect_and_update(canal_scores, 328, 1000)
    detect_and_update(canal_scores, 337, 100)
    detect_and_update(canal_scores, 343, 10)
    detect_and_update(canal_scores, 349, 1)

    detect_and_update(canal_scores, 332, 100)
    detect_and_update(canal_scores, 338, 10)
    detect_and_update(canal_scores, 333, 1)

    nickname_images = capture[:, 6:76].reshape(17, 24, -1)[:, 7:20]

    if not os.path.exists('nick_hash'):
        os.makedirs('nick_hash')

    page_query = 0
    page_updated_query = 0
    for nickname_image, mission, canal, warn in zip(nickname_images, mission_scores, canal_scores, flag_scores):
        nickname_hash = hashlib.sha256(str(nickname_image.tobytes()).encode()).hexdigest()
        if nickname_hash != BLANK_HASH:
            page_query += 1
            if nickname_hash not in written_nicknames:
                cv2.imwrite(f'nick_hash/{nickname_hash}.png', 255 - nickname_image * 255)
                written_nicknames[nickname_hash] = (mission, canal, warn)
                page_updated_query += 1

    if 0 < page_updated_query < page_query or 0 < page_updated_query < is_last_page_captured:
        print("\r============================경고============================")
        print("캡처된 페이지 중 일부 닉네임만 변동된 경우가 감지되었습니다.")
        print("이는 보통 캡처 도중 마우스 커서가 닉네임을 가린 경우 발생합니다.")
        print("마우스가 닉네임을 가리지 않게 조심해 주세요. [Enter]를 입력해 다시 스캔해 주세요.")
        print("============================================================")
        written_nicknames = {}
        scanned_query = 0
        is_last_page_captured = 0
        waiting_second = 10
        input("입력 대기 중...")
        continue

    if 0 < page_updated_query < 17:  # 17 is maximum query in a page
        if is_last_page_captured > 0:
            print("\r============================경고============================")
            print("마지막 페이지(17개 미만)로 인식된 페이지가 2회 이상 감지되었습니다.")
            print("이는 길드 창이 잘렸거나, 매우 드문 경우 닉네임 해시가 충돌하는 경우 발생합니다.")
            print("가린 창을 치운 후, [Enter]를 입력해 다시 스캔해 주세요.")
            print("============================================================")
            written_nicknames = {}
            scanned_query = 0
            is_last_page_captured = 0
            waiting_second = 10
            input("입력 대기 중...")
            continue
        else:
            is_last_page_captured = page_updated_query

    if page_updated_query:  # if updated
        scanned_query += page_updated_query
        print(f"\r{page_updated_query}명 추가, 누적 {scanned_query}명 기록")
        start_time = time.time()
    elif (time.time() - start_time) > waiting_second:
        print('\r시간이 초과되어 스캔을 종료합니다.', end='\n\n')
        break
    elif scanned_query > 187:  # 17*11=187, 12th page is last page.
        print('\r모든 페이지를 스캔하여 스캔을 종료합니다.', end='\n\n')
        break
    else:
        if scanned_query > 17:
            waiting_second = 5
        print(f"\r스크롤을 넘겨주세요. 새로운 페이지가 {start_time + waiting_second - time.time():.1f}초간 스캔되지 않으면 스캔을 종료합니다.", end='')
# scanning end


# write to xlsx file
# create for load first page
try:  # if already workbook exist
    wb = load_workbook('result.xlsx')
    w0 = wb.worksheets[0]
    hash2nick = {hash_: name
                 for hash_, name in w0.iter_rows(min_row=2, max_row=w0.max_row, min_col=2, max_col=3, values_only=True)}
except FileNotFoundError:  # create [image - hash - nickname] page
    wb = Workbook()
    w0 = wb.worksheets[0]
    w0.title = "닉네임 정보"
    w0.column_dimensions['B'].hidden = True
    w0.cell(row=1, column=1).value = '닉네임 이미지'
    w0.cell(row=1, column=2).value = '닉네임 해시'
    w0.cell(row=1, column=3).value = '닉네임(수동입력)'
    hash2nick = {}

# create today's page
ws = wb.create_sheet(title=datetime.now().strftime("%y%m%d_%H%M"))
ws.cell(row=1, column=1).value = '닉네임 해시'
ws.column_dimensions['A'].hidden = True
ws.cell(row=1, column=2).value = '닉네임(자동으로 채워짐)'
ws.cell(row=1, column=3).value = '주간미션'
ws.cell(row=1, column=4).value = '지하 수로'
ws.cell(row=1, column=5).value = '플래그 레이스'

naive_query = 0
for current_row, (hash_, (m, s, f)) in enumerate(written_nicknames.items(), start=2):
    ws.cell(row=current_row, column=1).value = hash_
    ws.cell(row=current_row, column=3).value = m
    ws.cell(row=current_row, column=4).value = s
    ws.cell(row=current_row, column=5).value = f
    if hash_ in hash2nick:
        if (nickname_written := hash2nick.pop(hash_)) is not None:
            ws.cell(row=current_row, column=2).value = nickname_written
            continue
    else:
        naive_query += 1

    formula = f"VLOOKUP(A{current_row},'닉네임 정보'!B:C,2,FALSE)"
    ws.cell(row=current_row, column=2).value = f'=IF({formula}<>0,{formula},"???")'

    max_r = w0.max_row + 1
    w0.add_image(Image(f'nick_hash/{hash_}.png'), anchor=f'A{max_r}')
    w0.cell(row=max_r, column=2).value = hash_

warn = False
if len(hash2nick) > 0:
    print("============================경고============================")
    print(f"기존에 존재한 닉네임 중 총 {len(hash2nick)}개가 스캔되지 않았습니다.")
    print(f"스캔 도중 마우스가 닉네임을 가렸거나, 닉네임을 변경했거나, 길드를 탈퇴하였을 수 있습니다.")
    print("누락된 닉네임 중, 수기로 기록되어 있던 닉네임은 다음과 같습니다.")
    print([v for v in list(hash2nick.values()) if v])
    warn = True
if naive_query > 0:
    if not warn:
        print("============================경고============================")
    print(f"기존에 없던 {naive_query}개의 닉네임이 추가되었습니다.")
    print("프로그램 종료 후 [닉네임 정보] 시트에서 닉네임을 수동으로 입력해 주세요.")
    print("[닉네임 정보] 시트에 수기로 기록하지 않은 닉네임은 기록 시트에서 ???로 표시됩니다.")
    warn = True

if warn:
    input("변동사항을 적용하려면 [Enter]를 입력하세요.\n============================================================\n")
else:
    print(f"총 {scanned_query}개의 닉네임이 스캔되었으며, 기존 닉네임 중 변동된 닉네임은 없습니다.")


while True:
    try:
        wb.save('result.xlsx')
        input("result.xlsx 파일이 저장되었습니다. [Enter]를 입력하거나 창을 닫아 종료하세요.")
        break
    except PermissionError:
        input("권한 오류가 발생하였습니다. [Enter]를 입력해 재시도합니다. 주로 result.xlsx가 열려 있는 경우 발생합니다.")
